# import openpyxl
# cate=openpyxl.load_workbook('test_case_api.xlsx')
# sheet=cate['register']
# danyuan=sheet.cell(row=3,column=3)
# print(danyuan.value)

#
# import openpyxl
# cate1=openpyxl.load_workbook('test_case_api.xlsx')
# sheet1=cate1['register']
# danyuan1=sheet1.cell(row=5,column=3)
# print(danyuan1.value)
# danyuan1.value='手机号长度10位'
# cate1.save('test_case_api.xlsx')

# import openpyxl
# cata=openpyxl.load_workbook('test_case_api.xlsx')
# sheet=cata['register']
# danyuan=sheet.cell(row=3,column=3)
# print(danyuan.value)
# danyuan.value='成功不带用户名'
# cata.save('test_case_api.xlsx')
# print(danyuan.value)




# def inp():
#     import openpyxl
#     cate=openpyxl.load_workbook('test_case_api.xlsx')
#     sheet=cate['register']
#     list_1=[]
#     max_row=sheet.max_row
#     for i in range(2,max_row+1):
#         dict_1=dict(
#         url=sheet.cell(row=i,column=5).value,
#         data=sheet.cell(row=i,column=6).value,
#         expected=sheet.cell(row=i,column=7).value)
#         list_1.append(dict_1)
#     return list_1
# cate=inp()
# print(cate[0]["url"])









import requests
import openpyxl
cate=openpyxl.load_workbook('test_case_api.xlsx')
sheet=cate['register']
list_1=[]
row_max=sheet.max_row
for i in range(2,row_max+1):
    dict_1=dict(
    url=sheet.cell(row=i,column=5).value,
    data=sheet.cell(row=i,column=6).value,
    expend=sheet.cell(row=i,column=7).value)
    list_1.append(dict_1)
# print(list_1)
for j in list_1:
    # print(j)
    expend=j['expend']
    expend=eval(expend)
    expend_msg=expend['msg']
    request_body=j['data']
    request_body=eval(request_body)
    # print(request_body)
    header = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    req = requests.post(url=j['url'],json=request_body,headers=header)
    req_c=req.json(encoding='utf-8')
    print(req_c)
    real_msg=req.json()['msg']
    if expend_msg == real_msg:
        print('pass')
    else:
        print('ng')
    print('*'*60)


# print(list_1)
# for j in re:
#     request=j
#     # print(j)
#     print(request['data'])
#     header={"X-Lemonban-Media-Type": "lemonban.v2","Content-Type": "application/json"}
#     req = requests.post(url=request['url'],json=request['data'],headers=header)
#     reqq = req.json()
#     # print(reqq)
#     # print('************************************************************************************')

