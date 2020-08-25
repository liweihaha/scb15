import openpyxl
import requests

# biaoge = openpyxl.load_workbook('test_case_api.xlsx') #打开excel表格加载到内存中
# sheet = biaoge['register']  #打开表格中具体的表单
# danyuange = sheet.cell(row=2, column=3)  #找到具体单元格
# print(danyuange.value)
#
# danyuange.value='哈哈哈'
# biaoge.save('test_case_api.xlsx') #重新赋值并保存


# 读取excel测试用例
def duqu(filename, sheet):
    biaoge = openpyxl.load_workbook(filename)
    sheet1 = biaoge[sheet]
    max_row = sheet1.max_row
    list1 = []
    for i in range(2, max_row+1):
        dict1 = dict(
        id=sheet1.cell(row=i, column=1).value,
        url = sheet1.cell(row=i, column=5).value,
        data = sheet1.cell(row=i, column=6).value,
        expect = sheet1.cell(row=i, column=7).value)
        list1.append(dict1)
    return list1


# 发送请求
def fasong(url, body):
    requests_header = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    res = requests.post(url=url, json=body, headers=requests_header)
    res_log = res.json()
    return res_log


# 把结果写到excel中
def xieru(filename, sheetname, row, column, final_result):
    biaoge = openpyxl.load_workbook(filename)
    sheet1 = biaoge[sheetname]
    sheet1.cell(row=row, column=column).value = final_result
    biaoge.save(filename)


# 执行测试，断言
def zhixing(filename,sheetname):
    res = duqu(filename, sheetname)
    for testcase in res:     # 取出一条一条的测试用例
        case_id = testcase.get('id')    # 字典取值或者value，取出id
        url = testcase.get('url')     # 取出url
        data = testcase.get('data')    # 取出 data 从excel取出来数据都是str
        data = eval(data)  # 运行被字符串包裹的python表达式，转换成字典格式    -- 用eval()把引号去掉
        expect = testcase.get('expect')  # 取出expect
        expect = eval(expect)   # 把字符串转换成字典
        expect_msg = expect.get('msg')   # 从预期结果的字典里把msg取出来
        # print(case_id, url, data, expect)
        res_1 = fasong(url=url, body=data)   # 调用发送请求的函数，并传入参数
        # print(res_1)
        real_msg = res_1.get('msg')
        if real_msg == expect_msg:
            final_res = '通过'
        else:
            final_res = '不通过，有bug'
        xieru(filename, sheetname, case_id+1, 8, final_res)


zhixing('test_case_api.xlsx', 'login')
zhixing('test_case_api.xlsx', 'register')